#!/usr/bin/env python3
"""
PROJECT CONTEXT -- Indexation + RAG Retrieval for Projects (S58)
================================================================

Provides two main classes:
- ProjectIndexer: indexes project files into per-project ChromaDB collections
  using the existing RAG chunker/embeddings infrastructure.
- ProjectContextBuilder: retrieves relevant chunks from a project's collection
  and builds a context string within a token budget for injection into LLM calls.

Graceful degradation: if ChromaDB or Ollama embeddings are unavailable,
indexation is skipped but system_instructions are still injected.

Author: Leon
"""

import logging
import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# =============================================================================
# CONDITIONAL IMPORTS
# =============================================================================

# ChromaDB
try:
    import chromadb
    from chromadb.config import Settings as ChromaSettings
    CHROMADB_AVAILABLE = True
except ImportError:
    CHROMADB_AVAILABLE = False
    chromadb = None
    ChromaSettings = None

# RAG chunkers
try:
    from opti_oignon.rag.chunkers import Chunk, get_chunker
    CHUNKERS_AVAILABLE = True
except ImportError:
    CHUNKERS_AVAILABLE = False
    get_chunker = None
    Chunk = None

# RAG embeddings
try:
    from opti_oignon.rag.config import get_config as get_rag_config
    from opti_oignon.rag.embeddings import OllamaEmbeddings
    EMBEDDINGS_AVAILABLE = True
except ImportError:
    EMBEDDINGS_AVAILABLE = False
    OllamaEmbeddings = None
    get_rag_config = None

# ProjectStore for updating file records
try:
    from opti_oignon.projects import ProjectFile, project_store
    PROJECTS_AVAILABLE = True
except ImportError:
    PROJECTS_AVAILABLE = False
    project_store = None
    ProjectFile = None

# Config
try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

# =============================================================================
# CONSTANTS
# =============================================================================

_CONFIG_DIR = Path(__file__).parent / "config"
_DATA_DIR = Path(__file__).parent / "data"
_DEFAULT_CHROMA_BASE = _DATA_DIR / "project_chroma"

# Collection name pattern for per-project collections
_COLLECTION_PREFIX = "project_"

# File type mapping from project categories to RAG chunker types
_PROJECT_TYPE_TO_RAG_TYPE: dict[str, str] = {
    "text": "text",
    "code": "python",
    "data": "csv",
    "document": "text",
    "image": None,
    "archive": None,
    "unknown": "text",
}

# Extension-specific overrides for more precise chunker selection
_EXT_TO_RAG_TYPE: dict[str, str] = {
    ".py": "python",
    ".r": "r",
    ".R": "r",
    ".rmd": "rmarkdown",
    ".Rmd": "rmarkdown",
    ".md": "markdown",
    ".markdown": "markdown",
    ".js": "javascript",
    ".ts": "typescript",
    ".sh": "shell",
    ".bash": "shell",
    ".sql": "sql",
    ".csv": "csv",
    ".tsv": "csv",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".json": "json",
    ".toml": "toml",
    ".html": "html",
    ".css": "css",
    ".txt": "text",
    ".rst": "text",
    ".log": "text",
    ".xml": "text",
    ".tex": "text",
}

# Default config values
_DEFAULT_CONTEXT_CONFIG = {
    "chroma_base_path": "project_chroma",
    "summary_max_tokens": 200,
    "key_terms_max": 20,
    "context_header": "--- Project Context ---",
    "context_footer": "--- End Project Context ---",
    "max_chunks_per_query": 10,
    "min_relevance_score": 0.25,
}


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class IndexResult:
    """Result of indexing a single project file."""

    file_id: str = ""
    filename: str = ""
    success: bool = False
    chunk_count: int = 0
    summary: str = ""
    key_terms: list[str] = field(default_factory=list)
    error: str = ""


@dataclass
class RetrievedChunk:
    """A chunk retrieved from the project's ChromaDB collection."""

    content: str = ""
    score: float = 0.0
    source_file: str = ""
    section_name: str = ""
    chunk_index: int = 0
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class ProjectContext:
    """Built context ready for injection into an LLM call."""

    context_text: str = ""
    system_instructions: str = ""
    chunks_used: int = 0
    total_tokens_estimate: int = 0
    source_files: list[str] = field(default_factory=list)
    project_id: str = ""
    project_name: str = ""


# =============================================================================
# HELPERS
# =============================================================================

def _estimate_tokens(text: str) -> int:
    """Estimate token count (approx 1 token per 4 characters)."""
    return len(text) // 4


def _l2_normalize(vec: Any) -> list[float]:
    """L2-normalize an embedding vector (PCX-01, S192).

    The per-project collections are created in ChromaDB's default l2 space
    and the retrieval score is computed as 1 - distance/2, which is only
    correct for unit-normalized vectors (squared L2 = 2 - 2*cos). The
    modern Ollama /api/embed endpoint normalizes server-side, so this is
    an idempotent no-op there; on the legacy /api/embeddings fallback
    (RST-01) vectors are unnormalized, every score collapsed to 0 below
    min_relevance_score, and project retrieval silently returned nothing.
    Normalizing at both seams (index and query) makes the score correct in
    both modes, with no migration needed for modern-endpoint data. The
    cosine-space alignment with rag_store (plus reindex) is recorded for
    the RAG cycle (PCX-03).
    """
    norm = math.sqrt(sum(x * x for x in vec))
    if norm <= 0.0:
        return list(vec)
    return [x / norm for x in vec]


def _get_rag_file_type(filename: str, project_file_type: str) -> str | None:
    """Map a project filename to a RAG chunker file type.

    Uses extension-specific overrides first, then falls back to
    the project category mapping.

    Args:
        filename: The file name (with extension).
        project_file_type: The project-level file type category.

    Returns:
        RAG file type string for get_chunker(), or None if not indexable.
    """
    ext = Path(filename).suffix.lower()
    if ext in _EXT_TO_RAG_TYPE:
        return _EXT_TO_RAG_TYPE[ext]
    return _PROJECT_TYPE_TO_RAG_TYPE.get(project_file_type, "text")


def _extract_key_terms(text: str, max_terms: int = 20) -> list[str]:
    """Extract key terms from text using simple frequency analysis.

    Focuses on meaningful tokens: filters out short words, common stop
    words, and returns the most frequent distinctive terms.

    Args:
        text: Source text to analyze.
        max_terms: Maximum number of terms to return.

    Returns:
        List of key term strings.
    """
    # Common English/French stop words to exclude
    stop_words = {
        "the", "a", "an", "is", "are", "was", "were", "be", "been", "being",
        "have", "has", "had", "do", "does", "did", "will", "would", "could",
        "should", "may", "might", "shall", "can", "need", "must", "to", "of",
        "in", "for", "on", "with", "at", "by", "from", "as", "into", "through",
        "during", "before", "after", "above", "below", "between", "out", "off",
        "over", "under", "again", "further", "then", "once", "here", "there",
        "when", "where", "why", "how", "all", "each", "every", "both", "few",
        "more", "most", "other", "some", "such", "no", "not", "only", "own",
        "same", "so", "than", "too", "very", "just", "because", "but", "and",
        "or", "if", "while", "about", "up", "that", "this", "these", "those",
        "it", "its", "he", "she", "they", "them", "we", "you", "i", "my",
        "your", "his", "her", "our", "their", "what", "which", "who", "whom",
        "le", "la", "les", "un", "une", "des", "de", "du", "et", "en", "est",
        "que", "qui", "dans", "pour", "sur", "avec", "pas", "plus", "par",
        "ce", "se", "son", "sa", "ses", "au", "aux", "ne", "ou", "mais",
        "import", "def", "class", "return", "self", "none", "true",
        "false", "print", "str", "int", "float", "list", "dict",
    }

    # Tokenize: extract word-like tokens
    words = re.findall(r"[a-zA-Z_][a-zA-Z0-9_]{2,}", text)
    words_lower = [w.lower() for w in words]

    # Count frequencies, excluding stop words
    freq: dict[str, int] = {}
    for w in words_lower:
        if w not in stop_words and len(w) >= 3:
            freq[w] = freq.get(w, 0) + 1

    # Sort by frequency descending, take top terms
    sorted_terms = sorted(freq.items(), key=lambda x: x[1], reverse=True)
    return [term for term, _ in sorted_terms[:max_terms]]


def _generate_summary(text: str, max_tokens: int = 200) -> str:
    """Generate a simple extractive summary from text.

    Takes the first few sentences up to the token budget.
    A proper LLM-based summary could be added later.

    Args:
        text: Source text.
        max_tokens: Approximate max tokens for the summary.

    Returns:
        Summary string.
    """
    max_chars = max_tokens * 4
    # Take the beginning of the text, trying to break at sentence boundaries
    if len(text) <= max_chars:
        return text.strip()

    # Find a sentence break near the limit
    truncated = text[:max_chars]
    last_period = truncated.rfind(".")
    last_newline = truncated.rfind("\n")
    break_point = max(last_period, last_newline)
    if break_point > max_chars // 2:
        return truncated[:break_point + 1].strip()
    return truncated.strip() + "..."


def _load_context_config() -> dict[str, Any]:
    """Load project context configuration from projects.yaml.

    Returns:
        Dict of context-related config values.
    """
    config = dict(_DEFAULT_CONTEXT_CONFIG)
    config_path = _CONFIG_DIR / "projects.yaml"
    if not YAML_AVAILABLE or not config_path.exists():
        return config

    try:
        with open(config_path) as f:
            raw = yaml.safe_load(f) or {}
        ctx_cfg = raw.get("projects", {}).get("context", {})
        config.update(ctx_cfg)
    except Exception as e:
        logger.warning("Failed to load project context config: %s", e)

    return config


# =============================================================================
# PROJECT INDEXER
# =============================================================================

class ProjectIndexer:
    """Indexes project files into per-project ChromaDB collections.

    Each project gets its own ChromaDB collection named 'project_{id}'.
    Uses the existing RAG chunkers and OllamaEmbeddings for processing.
    After indexing, updates the ProjectFile record in SQLite with
    indexed=True, chunk_count, summary, and key_terms.

    Graceful degradation: if ChromaDB or embeddings are unavailable,
    returns an error IndexResult without crashing.
    """

    def __init__(
        self,
        chroma_base: Path | None = None,
        store: Any | None = None,
    ):
        """Initialize the project indexer.

        Args:
            chroma_base: Base directory for per-project ChromaDB storage.
            store: ProjectStore instance (uses global singleton if None).
        """
        self._config = _load_context_config()
        self._chroma_base = chroma_base or _DEFAULT_CHROMA_BASE
        self._chroma_base.mkdir(parents=True, exist_ok=True)
        self._store = store if store is not None else project_store
        self._embedder: Any | None = None

        # Lazy-init embedder on first use
        self._embedder_checked = False

    @property
    def available(self) -> bool:
        """Whether indexation is possible (ChromaDB + chunkers + embeddings)."""
        return CHROMADB_AVAILABLE and CHUNKERS_AVAILABLE and EMBEDDINGS_AVAILABLE

    def _get_embedder(self) -> Any | None:
        """Get or create the OllamaEmbeddings instance (lazy init)."""
        if self._embedder is not None:
            return self._embedder
        if not EMBEDDINGS_AVAILABLE:
            return None
        if self._embedder_checked:
            return self._embedder

        self._embedder_checked = True
        try:
            rag_config = get_rag_config()
            self._embedder = OllamaEmbeddings(rag_config.embedding)
            return self._embedder
        except Exception as e:
            logger.warning("Failed to initialize embedder: %s", e)
            return None

    def _get_collection(self, project_id: str) -> Any | None:
        """Get or create the ChromaDB collection for a project.

        Args:
            project_id: The project identifier.

        Returns:
            ChromaDB collection, or None if unavailable.
        """
        if not CHROMADB_AVAILABLE:
            return None

        try:
            client = chromadb.PersistentClient(
                path=str(self._chroma_base),
                settings=ChromaSettings(anonymized_telemetry=False),
            )
            collection_name = f"{_COLLECTION_PREFIX}{project_id}"
            return client.get_or_create_collection(
                name=collection_name,
                metadata={"project_id": project_id},
            )
        except Exception as e:
            logger.error("Failed to get ChromaDB collection for project %s: %s", project_id, e)
            return None

    def _read_file_content(self, file_path: str, filename: str) -> str | None:
        """Read file content with encoding fallback.

        Supports text-based files. Binary formats (PDF, Excel, DOCX)
        are read via the existing RAG indexer helpers if available.

        Args:
            file_path: Absolute path to the file.
            filename: Original filename for type detection.

        Returns:
            Text content, or None if unreadable.
        """
        fp = Path(file_path)
        if not fp.exists():
            return None

        ext = fp.suffix.lower()

        # Binary format extraction
        if ext == ".pdf":
            try:
                from pypdf import PdfReader
                reader = PdfReader(str(fp))
                parts = []
                for i, page in enumerate(reader.pages):
                    t = page.extract_text()
                    if t:
                        parts.append(f"[Page {i + 1}]\n{t}")
                return "\n\n".join(parts) if parts else None
            except Exception as e:
                logger.warning("PDF extraction failed for %s: %s", filename, e)
                return None

        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(fp), data_only=True)
                parts = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        row_str = [str(c) if c is not None else "" for c in row]
                        if any(s.strip() for s in row_str):
                            rows.append(",".join(row_str))
                    if rows:
                        parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
                return "\n\n".join(parts) if parts else None
            except Exception as e:
                logger.warning("Excel extraction failed for %s: %s", filename, e)
                return None

        if ext == ".docx":
            try:
                import docx
                doc = docx.Document(str(fp))
                parts = [p.text for p in doc.paragraphs if p.text.strip()]
                return "\n\n".join(parts) if parts else None
            except Exception as e:
                logger.warning("DOCX extraction failed for %s: %s", filename, e)
                return None

        # Text-based files: try multiple encodings
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                with open(fp, encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
            except Exception as e:
                logger.warning("Read error for %s: %s", filename, e)
                return None

        return None

    def index_file(
        self,
        project_id: str,
        file_id: str,
    ) -> IndexResult:
        """Index a single project file into ChromaDB.

        Reads the file, chunks it, generates embeddings, stores in
        the per-project collection, and updates the SQLite record.

        Args:
            project_id: The project ID.
            file_id: The project file ID.

        Returns:
            IndexResult with success/failure details.
        """
        result = IndexResult(file_id=file_id)

        # Check dependencies
        if not self.available:
            missing = []
            if not CHROMADB_AVAILABLE:
                missing.append("chromadb")
            if not CHUNKERS_AVAILABLE:
                missing.append("chunkers")
            if not EMBEDDINGS_AVAILABLE:
                missing.append("embeddings")
            result.error = f"Missing dependencies: {', '.join(missing)}"
            return result

        # Get the file record
        if self._store is None:
            result.error = "ProjectStore not available"
            return result

        pf = self._store.get_file(file_id)
        if pf is None:
            result.error = f"File {file_id} not found"
            return result

        result.filename = pf.filename

        # Determine RAG file type
        rag_type = _get_rag_file_type(pf.filename, pf.file_type)
        if rag_type is None:
            result.error = f"File type '{pf.file_type}' not indexable (binary)"
            return result

        # Read file content
        content = self._read_file_content(pf.file_path, pf.filename)
        if not content or len(content.strip()) < 10:
            result.error = "File is empty or too short to index"
            return result

        # Get ChromaDB collection
        collection = self._get_collection(project_id)
        if collection is None:
            result.error = "ChromaDB collection unavailable"
            return result

        # Get embedder
        embedder = self._get_embedder()
        if embedder is None:
            result.error = "Embedding model unavailable"
            return result

        # Remove old chunks for this file (reindex safe)
        try:
            existing = collection.get(where={"source_file": pf.file_path})
            if existing and existing["ids"]:
                collection.delete(ids=existing["ids"])
        except Exception:
            pass

        # Chunk the content
        try:
            chunker = get_chunker(rag_type)
            chunks = chunker.chunk(content, pf.file_path, rag_type)
        except Exception as e:
            result.error = f"Chunking failed: {e}"
            return result

        if not chunks:
            result.error = "No chunks produced"
            return result

        # Generate embeddings
        try:
            chunk_texts = [c.content for c in chunks]
            embeddings = embedder.embed(chunk_texts, show_progress=False)
        except Exception as e:
            result.error = f"Embedding failed: {e}"
            return result

        # Filter valid embeddings
        valid_chunks = []
        valid_embeddings = []
        for chunk, emb in zip(chunks, embeddings):
            if emb is not None:
                valid_chunks.append(chunk)
                valid_embeddings.append(emb)

        if not valid_chunks:
            result.error = "No valid embeddings produced"
            return result

        # PCX-01 (S192): normalize before storing (see _l2_normalize).
        valid_embeddings = [_l2_normalize(e) for e in valid_embeddings]

        # Store in ChromaDB
        try:
            # Use project-scoped chunk IDs to avoid collisions
            chunk_ids = [
                f"{project_id}::{pf.file_path}::{c.chunk_index}"
                for c in valid_chunks
            ]
            metadatas = [c.metadata for c in valid_chunks]
            collection.add(
                ids=chunk_ids,
                embeddings=valid_embeddings,
                documents=[c.content for c in valid_chunks],
                metadatas=metadatas,
            )
        except Exception as e:
            result.error = f"ChromaDB insertion failed: {e}"
            return result

        # Extract summary and key terms
        summary = _generate_summary(content, self._config.get("summary_max_tokens", 200))
        key_terms = _extract_key_terms(content, self._config.get("key_terms_max", 20))

        result.success = True
        result.chunk_count = len(valid_chunks)
        result.summary = summary
        result.key_terms = key_terms

        # Update the SQLite record
        try:
            self._update_file_record(
                file_id=file_id,
                indexed=True,
                chunk_count=len(valid_chunks),
                summary=summary,
                key_terms=key_terms,
            )
        except Exception as e:
            logger.warning("Failed to update file record after indexing: %s", e)

        logger.info(
            "Indexed file '%s' for project %s: %d chunks, %d key terms",
            pf.filename, project_id, len(valid_chunks), len(key_terms),
        )
        return result

    def remove_file_from_index(self, project_id: str, file_id: str) -> bool:
        """Remove a file's chunks from the project's ChromaDB collection.

        Args:
            project_id: The project ID.
            file_id: The project file ID.

        Returns:
            True if removal succeeded.
        """
        if not CHROMADB_AVAILABLE:
            return False

        if self._store is None:
            return False

        pf = self._store.get_file(file_id)
        if pf is None:
            return False

        collection = self._get_collection(project_id)
        if collection is None:
            return False

        try:
            existing = collection.get(where={"source_file": pf.file_path})
            if existing and existing["ids"]:
                collection.delete(ids=existing["ids"])
                logger.info(
                    "Removed %d chunks for file '%s' from project %s",
                    len(existing["ids"]), pf.filename, project_id,
                )
        except Exception as e:
            logger.warning("Failed to remove chunks from index: %s", e)
            return False

        # Reset the indexed fields
        try:
            self._update_file_record(
                file_id=file_id,
                indexed=False,
                chunk_count=0,
                summary="",
                key_terms=[],
            )
        except Exception:
            pass

        return True

    def reindex_project(self, project_id: str) -> list[IndexResult]:
        """Reindex all files in a project.

        Args:
            project_id: The project ID.

        Returns:
            List of IndexResult for each file.
        """
        results = []

        if self._store is None:
            return results

        files = self._store.list_files(project_id)
        for pf in files:
            result = self.index_file(project_id, pf.id)
            results.append(result)

        return results

    def delete_project_index(self, project_id: str) -> bool:
        """Delete the entire ChromaDB collection for a project.

        Args:
            project_id: The project ID.

        Returns:
            True if deletion succeeded.
        """
        if not CHROMADB_AVAILABLE:
            return False

        try:
            client = chromadb.PersistentClient(
                path=str(self._chroma_base),
                settings=ChromaSettings(anonymized_telemetry=False),
            )
            collection_name = f"{_COLLECTION_PREFIX}{project_id}"
            client.delete_collection(collection_name)
            logger.info("Deleted ChromaDB collection for project %s", project_id)
            return True
        except Exception as e:
            logger.warning("Failed to delete project index: %s", e)
            return False

    def _update_file_record(
        self,
        file_id: str,
        indexed: bool,
        chunk_count: int,
        summary: str,
        key_terms: list[str],
    ) -> None:
        """Update the project_files SQLite record with indexation results.

        Args:
            file_id: The file ID to update.
            indexed: Whether the file is now indexed.
            chunk_count: Number of chunks produced.
            summary: File summary text.
            key_terms: List of extracted key terms.
        """
        if self._store is None:
            return

        import json
        conn = self._store._get_conn()
        try:
            conn.execute(
                """
                UPDATE project_files
                SET indexed = ?, chunk_count = ?, summary = ?,
                    key_terms = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    int(indexed),
                    chunk_count,
                    summary,
                    json.dumps(key_terms),
                    _iso_now(),
                    file_id,
                ),
            )
            conn.commit()
        finally:
            conn.close()


# =============================================================================
# PROJECT CONTEXT BUILDER
# =============================================================================

class ProjectContextBuilder:
    """Builds context from a project's indexed files for LLM injection.

    Retrieves relevant chunks from ChromaDB based on a query,
    assembles them within a token budget, and prepends the project's
    system_instructions.

    Token budget is a hard limit: the builder never exceeds
    context_budget_tokens from the project settings.
    """

    def __init__(
        self,
        chroma_base: Path | None = None,
        store: Any | None = None,
    ):
        """Initialize the context builder.

        Args:
            chroma_base: Base directory for ChromaDB storage.
            store: ProjectStore instance (uses global singleton if None).
        """
        self._config = _load_context_config()
        self._chroma_base = chroma_base or _DEFAULT_CHROMA_BASE
        self._store = store if store is not None else project_store
        self._embedder: Any | None = None
        self._embedder_checked = False

    @property
    def available(self) -> bool:
        """Whether RAG retrieval is possible."""
        return CHROMADB_AVAILABLE and EMBEDDINGS_AVAILABLE

    def _get_embedder(self) -> Any | None:
        """Get or create the OllamaEmbeddings instance."""
        if self._embedder is not None:
            return self._embedder
        if not EMBEDDINGS_AVAILABLE:
            return None
        if self._embedder_checked:
            return self._embedder

        self._embedder_checked = True
        try:
            rag_config = get_rag_config()
            self._embedder = OllamaEmbeddings(rag_config.embedding)
            return self._embedder
        except Exception as e:
            logger.warning("Failed to initialize embedder: %s", e)
            return None

    def _get_collection(self, project_id: str) -> Any | None:
        """Get the ChromaDB collection for a project (read-only)."""
        if not CHROMADB_AVAILABLE:
            return None

        try:
            client = chromadb.PersistentClient(
                path=str(self._chroma_base),
                settings=ChromaSettings(anonymized_telemetry=False),
            )
            collection_name = f"{_COLLECTION_PREFIX}{project_id}"
            return client.get_collection(collection_name)
        except Exception:
            return None

    def retrieve_chunks(
        self,
        project_id: str,
        query: str,
        max_chunks: int | None = None,
        min_score: float | None = None,
    ) -> list[RetrievedChunk]:
        """Retrieve relevant chunks from a project's index.

        Args:
            project_id: The project ID.
            query: User query for semantic search.
            max_chunks: Max number of chunks to return.
            min_score: Minimum relevance score (0-1).

        Returns:
            List of RetrievedChunk sorted by relevance (best first).
        """
        if not self.available or not query.strip():
            return []

        max_chunks = max_chunks or self._config.get("max_chunks_per_query", 10)
        min_score = min_score if min_score is not None else self._config.get("min_relevance_score", 0.25)

        collection = self._get_collection(project_id)
        if collection is None:
            return []

        # Check collection has documents
        try:
            count = collection.count()
            if count == 0:
                return []
        except Exception:
            return []

        # Get query embedding
        embedder = self._get_embedder()
        if embedder is None:
            return []

        try:
            query_embedding = embedder.embed_single(query)
            if query_embedding is None:
                return []
            # PCX-01 (S192): normalize the query like the stored vectors.
            query_embedding = _l2_normalize(query_embedding)
        except Exception as e:
            logger.warning("Query embedding failed: %s", e)
            return []

        # Query ChromaDB
        try:
            results = collection.query(
                query_embeddings=[query_embedding],
                n_results=min(max_chunks * 2, 20),
                include=["documents", "metadatas", "distances"],
            )
        except Exception as e:
            logger.warning("ChromaDB query failed: %s", e)
            return []

        if not results or not results.get("documents"):
            return []

        documents = results["documents"][0]
        metadatas = results["metadatas"][0]
        distances = results["distances"][0]

        chunks = []
        for doc, meta, dist in zip(documents, metadatas, distances):
            # Convert distance to similarity score
            score = max(0, 1 - dist / 2)
            if score < min_score:
                continue

            chunks.append(RetrievedChunk(
                content=doc,
                score=score,
                source_file=meta.get("source_file", ""),
                section_name=meta.get("section_name", ""),
                chunk_index=meta.get("chunk_index", 0),
                metadata=meta,
            ))

        # Sort by score descending, limit
        chunks.sort(key=lambda c: c.score, reverse=True)
        return chunks[:max_chunks]

    def build_context(
        self,
        project_id: str,
        query: str,
        budget_tokens: int | None = None,
    ) -> ProjectContext:
        """Build a context block for injection into the LLM prompt.

        Retrieves relevant chunks, assembles them within the token
        budget, and prepends the project's system_instructions.

        Args:
            project_id: The project ID.
            query: User query for semantic retrieval.
            budget_tokens: Max tokens for the entire context block.
                Uses project setting if None.

        Returns:
            ProjectContext ready for injection.
        """
        ctx = ProjectContext(project_id=project_id)

        if self._store is None:
            return ctx

        # Get project details
        project = self._store.get_project(project_id)
        if project is None:
            return ctx

        ctx.project_name = project.name
        ctx.system_instructions = project.system_instructions or ""

        # Determine token budget
        if budget_tokens is None:
            budget_tokens = project.settings.get("context_budget_tokens", 4096)

        # Reserve tokens for system_instructions
        instructions_tokens = _estimate_tokens(ctx.system_instructions)
        header = self._config.get("context_header", "--- Project Context ---")
        footer = self._config.get("context_footer", "--- End Project Context ---")
        overhead_tokens = _estimate_tokens(header + footer) + 20
        available_for_chunks = budget_tokens - instructions_tokens - overhead_tokens

        if available_for_chunks <= 0:
            # Only system_instructions fit
            ctx.context_text = ctx.system_instructions
            ctx.total_tokens_estimate = instructions_tokens
            return ctx

        # Retrieve relevant chunks
        chunks = self.retrieve_chunks(project_id, query)

        # Assemble chunks within budget
        selected_parts = []
        total_chunk_tokens = 0
        source_files = set()

        for chunk in chunks:
            chunk_tokens = _estimate_tokens(chunk.content)
            if total_chunk_tokens + chunk_tokens > available_for_chunks:
                # Try to fit a truncated version
                remaining = available_for_chunks - total_chunk_tokens
                if remaining > 50:
                    truncated = chunk.content[:remaining * 4]
                    selected_parts.append(
                        f"[{Path(chunk.source_file).name}] {truncated}..."
                    )
                    source_files.add(Path(chunk.source_file).name)
                break

            source_label = Path(chunk.source_file).name
            if chunk.section_name:
                source_label += f" ({chunk.section_name})"
            selected_parts.append(f"[{source_label}]\n{chunk.content}")
            total_chunk_tokens += chunk_tokens
            source_files.add(Path(chunk.source_file).name)

        ctx.chunks_used = len(selected_parts)
        ctx.source_files = sorted(source_files)

        # Build final context text
        parts = []
        if ctx.system_instructions:
            parts.append(ctx.system_instructions)
        if selected_parts:
            parts.append(header)
            parts.extend(selected_parts)
            parts.append(footer)

        ctx.context_text = "\n\n".join(parts)
        ctx.total_tokens_estimate = _estimate_tokens(ctx.context_text)

        return ctx

    def build_system_instructions_only(self, project_id: str) -> ProjectContext:
        """Build a context with only system_instructions (no RAG).

        Used as fallback when ChromaDB is unavailable or when trigger
        detection determines the query is not project-relevant.

        Args:
            project_id: The project ID.

        Returns:
            ProjectContext with system_instructions only.
        """
        ctx = ProjectContext(project_id=project_id)

        if self._store is None:
            return ctx

        project = self._store.get_project(project_id)
        if project is None:
            return ctx

        ctx.project_name = project.name
        ctx.system_instructions = project.system_instructions or ""
        ctx.context_text = ctx.system_instructions
        ctx.total_tokens_estimate = _estimate_tokens(ctx.context_text)
        return ctx


# =============================================================================
# HELPER (timestamp)
# =============================================================================

def _iso_now() -> str:
    """Return current UTC time in ISO 8601 format."""
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


# =============================================================================
# MODULE-LEVEL SINGLETONS
# =============================================================================

project_indexer = ProjectIndexer()
project_context_builder = ProjectContextBuilder()
